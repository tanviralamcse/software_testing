import openpyxl

path = "/Users/tanviralam/xlsx1.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active

row = sheet.max_row
col = sheet.max_column

print(row)
print(col)

for r in range(1, row + 1):
    for c in range(1, col + 1):
        print(sheet.cell(row=r, column=c).value, end="          ")  # end=" " to print values on the same line
    print()